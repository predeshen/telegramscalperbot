"""Excel Reporter - Logs scan results to Excel and sends email reports."""
import logging
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl>=3.1.0")


logger = logging.getLogger(__name__)


class ExcelReporter:
    """Logs scan results to Excel file and sends periodic email reports."""
    
    def __init__(
        self,
        excel_file_path: str,
        smtp_config: dict,
        report_interval_seconds: int = 3600,
        initial_report_delay_seconds: int = 300,
        scanner_name: str = "Scanner"
    ):
        """
        Initialize Excel reporter.
        
        Args:
            excel_file_path: Path to Excel file
            smtp_config: SMTP configuration dict with keys: server, port, user, password, from_email, to_email, use_ssl
            report_interval_seconds: Interval between reports (default: 1 hour)
            initial_report_delay_seconds: Delay before first report (default: 5 min)
            scanner_name: Name of scanner for email subject
        """
        self.excel_file_path = Path(excel_file_path)
        self.smtp_config = smtp_config
        self.report_interval_seconds = report_interval_seconds
        self.initial_report_delay_seconds = initial_report_delay_seconds
        self.scanner_name = scanner_name
        
        # Thread control
        self.running = False
        self.file_lock = threading.Lock()
        self.initial_timer: Optional[threading.Timer] = None
        self.recurring_timer: Optional[threading.Timer] = None
        
        # Statistics
        self.scan_count = 0
        self.signal_count = 0
        self.long_signals = 0
        self.short_signals = 0
        self.start_time = datetime.now()
        self.last_report_time: Optional[datetime] = None
        
        # Excel logging disabled flag
        self.excel_disabled = False
        
        # Ensure directory exists
        self.excel_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Initialize Excel file
        self._initialize_excel_file()
        
        logger.info(f"ExcelReporter initialized: {self.excel_file_path}")
    
    def _initialize_excel_file(self) -> None:
        """Create Excel file with headers if it doesn't exist."""
        if self.excel_file_path.exists():
            logger.info(f"Using existing Excel file: {self.excel_file_path}")
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Scan Results"
            
            # Define headers
            headers = [
                'scan_id', 'timestamp', 'scanner', 'symbol', 'timeframe',
                'price', 'volume', 'ema_9', 'ema_21', 'ema_50', 'ema_100', 'ema_200',
                'rsi', 'atr', 'volume_ma', 'vwap', 'stoch_k', 'stoch_d',
                'signal_detected', 'signal_type', 'signal_entry', 'signal_sl',
                'signal_tp', 'signal_rr', 'signal_strategy'
            ]
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = cell.font.copy(bold=True)
            
            # Auto-fit columns
            for col_num in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 15
            
            # Enable filters
            ws.auto_filter.ref = ws.dimensions
            
            # Save file
            wb.save(self.excel_file_path)
            logger.info(f"Created new Excel file: {self.excel_file_path}")
            
        except Exception as e:
            logger.error(f"Failed to create Excel file: {e}")
            self.excel_disabled = True
    
    def log_scan_result(self, scan_data: Dict[str, Any]) -> bool:
        """
        Append scan result to Excel file.
        
        Args:
            scan_data: Dictionary containing scan result data with keys:
                - timestamp: datetime
                - scanner: str
                - symbol: str
                - timeframe: str
                - price: float
                - volume: float
                - indicators: dict (ema_9, ema_21, rsi, atr, etc.)
                - signal_detected: bool
                - signal_type: str (optional)
                - signal_details: dict (optional)
        
        Returns:
            True if successful, False otherwise
        """
        if self.excel_disabled:
            return False
        
        try:
            with self.file_lock:
                # Retry logic for locked files
                for attempt in range(3):
                    try:
                        wb = load_workbook(self.excel_file_path)
                        ws = wb.active
                        break
                    except Exception as e:
                        if attempt < 2:
                            logger.warning(f"Excel file locked, retrying... ({attempt + 1}/3)")
                            time.sleep(5)
                        else:
                            raise e
                
                # Increment scan count
                self.scan_count += 1
                
                # Extract data
                indicators = scan_data.get('indicators', {})
                signal_details = scan_data.get('signal_details', {})
                signal_detected = scan_data.get('signal_detected', False)
                
                # Update statistics
                if signal_detected:
                    self.signal_count += 1
                    signal_type = scan_data.get('signal_type', '')
                    if signal_type == 'LONG':
                        self.long_signals += 1
                    elif signal_type == 'SHORT':
                        self.short_signals += 1
                
                # Prepare row data
                row_data = [
                    self.scan_count,
                    scan_data.get('timestamp', datetime.now()).strftime('%Y-%m-%d %H:%M:%S'),
                    scan_data.get('scanner', self.scanner_name),
                    scan_data.get('symbol', ''),
                    scan_data.get('timeframe', ''),
                    scan_data.get('price', None),
                    scan_data.get('volume', None),
                    indicators.get('ema_9', None),
                    indicators.get('ema_21', None),
                    indicators.get('ema_50', None),
                    indicators.get('ema_100', None),
                    indicators.get('ema_200', None),
                    indicators.get('rsi', None),
                    indicators.get('atr', None),
                    indicators.get('volume_ma', None),
                    indicators.get('vwap', None),
                    indicators.get('stoch_k', None),
                    indicators.get('stoch_d', None),
                    signal_detected,
                    scan_data.get('signal_type', '') if signal_detected else '',
                    signal_details.get('entry_price', None) if signal_detected else None,
                    signal_details.get('stop_loss', None) if signal_detected else None,
                    signal_details.get('take_profit', None) if signal_detected else None,
                    signal_details.get('risk_reward', None) if signal_detected else None,
                    signal_details.get('strategy', '') if signal_detected else ''
                ]
                
                # Append row
                ws.append(row_data)
                
                # Save file
                wb.save(self.excel_file_path)
                
                return True
                
        except PermissionError:
            logger.error(f"Permission denied writing to Excel file: {self.excel_file_path}")
            self.excel_disabled = True
            return False
        except OSError as e:
            if "No space left on device" in str(e):
                logger.error("Disk full - cannot write to Excel file")
                self.excel_disabled = True
            else:
                logger.error(f"OS error writing to Excel file: {e}")
            return False
        except Exception as e:
            logger.error(f"Error logging scan result to Excel: {e}", exc_info=True)
            return False
    
    def _generate_email_body(self, is_initial: bool = False) -> str:
        """
        Generate HTML email body with summary statistics.
        
        Args:
            is_initial: Whether this is the initial startup report
        
        Returns:
            HTML email body string
        """
        now = datetime.now()
        report_period_start = self.last_report_time or self.start_time
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h2 {{ color: #2c3e50; }}
                .summary {{ background-color: #ecf0f1; padding: 15px; border-radius: 5px; }}
                .stat {{ margin: 8px 0; }}
                .label {{ font-weight: bold; }}
                .startup {{ background-color: #d5f4e6; padding: 10px; border-radius: 5px; margin-bottom: 15px; }}
            </style>
        </head>
        <body>
            <h2>📊 {self.scanner_name} Activity Report</h2>
        """
        
        if is_initial:
            html += f"""
            <div class="startup">
                <strong>🟢 Scanner Started</strong><br>
                Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S UTC')}<br>
                This is your initial verification report.
            </div>
            """
        
        html += f"""
            <div class="summary">
                <div class="stat"><span class="label">Scanner:</span> {self.scanner_name}</div>
                <div class="stat"><span class="label">Report Period:</span> {report_period_start.strftime('%Y-%m-%d %H:%M:%S')} to {now.strftime('%Y-%m-%d %H:%M:%S')} UTC</div>
                <div class="stat"><span class="label">Total Scans:</span> {self.scan_count}</div>
                <div class="stat"><span class="label">Signals Detected:</span> {self.signal_count}</div>
                <div class="stat"><span class="label">Signal Breakdown:</span> LONG: {self.long_signals}, SHORT: {self.short_signals}</div>
            </div>
            
            <p>📎 <strong>Attached:</strong> {self.excel_file_path.name}</p>
            
            <p style="color: #7f8c8d; font-size: 12px;">
                This is an automated report from your trading scanner.<br>
                Excel file contains detailed scan results with all indicators and signals.
            </p>
        </body>
        </html>
        """
        
        return html
    
    def _send_email_report(self, is_initial: bool = False) -> bool:
        """
        Send email report with Excel file attached.
        
        Args:
            is_initial: Whether this is the initial startup report
        
        Returns:
            True if successful, False otherwise
        """
        if not self.excel_file_path.exists():
            logger.warning("Excel file doesn't exist, skipping email report")
            return False
        
        try:
            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.smtp_config['from_email']
            msg['To'] = self.smtp_config['to_email']
            
            report_type = "Initial Verification" if is_initial else "Hourly"
            msg['Subject'] = f"[{self.scanner_name}] {report_type} Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            
            # Attach body
            body = self._generate_email_body(is_initial)
            msg.attach(MIMEText(body, 'html'))
            
            # Attach Excel file
            with open(self.excel_file_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename={self.excel_file_path.name}'
                )
                msg.attach(part)
            
            # Send email with retry logic
            for attempt in range(3):
                try:
                    if self.smtp_config.get('use_ssl', True):
                        server = smtplib.SMTP_SSL(
                            self.smtp_config['server'],
                            self.smtp_config['port'],
                            timeout=30
                        )
                    else:
                        server = smtplib.SMTP(
                            self.smtp_config['server'],
                            self.smtp_config['port'],
                            timeout=30
                        )
                        server.starttls()
                    
                    server.login(self.smtp_config['user'], self.smtp_config['password'])
                    server.send_message(msg)
                    server.quit()
                    
                    logger.info(f"Email report sent successfully ({report_type})")
                    self.last_report_time = datetime.now()
                    return True
                    
                except (smtplib.SMTPAuthenticationError, smtplib.SMTPException) as e:
                    logger.error(f"SMTP error sending email (attempt {attempt + 1}/3): {e}")
                    if attempt < 2:
                        time.sleep(2 ** attempt)  # Exponential backoff
                    else:
                        return False
                except Exception as e:
                    logger.error(f"Error sending email (attempt {attempt + 1}/3): {e}")
                    if attempt < 2:
                        time.sleep(2 ** attempt)
                    else:
                        return False
            
            return False
            
        except Exception as e:
            logger.error(f"Failed to send email report: {e}", exc_info=True)
            return False
    
    def _schedule_initial_report(self) -> None:
        """Schedule the initial report to be sent after delay."""
        if not self.running:
            return
        
        logger.info(f"Initial report scheduled in {self.initial_report_delay_seconds} seconds")
        
        def send_initial():
            if self.running:
                logger.info("Sending initial verification report...")
                self._send_email_report(is_initial=True)
                # Schedule recurring reports
                self._schedule_recurring_report()
        
        self.initial_timer = threading.Timer(self.initial_report_delay_seconds, send_initial)
        self.initial_timer.daemon = True
        self.initial_timer.start()
    
    def _schedule_recurring_report(self) -> None:
        """Schedule recurring hourly reports."""
        if not self.running:
            return
        
        def send_recurring():
            if self.running:
                logger.info("Sending scheduled report...")
                self._send_email_report(is_initial=False)
                # Schedule next report
                self._schedule_recurring_report()
        
        self.recurring_timer = threading.Timer(self.report_interval_seconds, send_recurring)
        self.recurring_timer.daemon = True
        self.recurring_timer.start()
    
    def start(self) -> None:
        """Start the reporter (schedules email reports)."""
        if self.running:
            logger.warning("ExcelReporter already running")
            return
        
        self.running = True
        logger.info("ExcelReporter started")
        
        # Schedule initial report
        self._schedule_initial_report()
    
    def stop(self) -> None:
        """Stop the reporter gracefully."""
        if not self.running:
            return
        
        logger.info("Stopping ExcelReporter...")
        self.running = False
        
        # Cancel timers
        if self.initial_timer:
            self.initial_timer.cancel()
        if self.recurring_timer:
            self.recurring_timer.cancel()
        
        logger.info("ExcelReporter stopped")
